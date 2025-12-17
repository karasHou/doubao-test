#!/usr/bin/env python3
import sys
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from fpdf import FPDF

def generate_excel_report(data, report_type):
    """生成 Excel 报表"""
    wb = Workbook()
    ws = wb.active

    # 设置标题
    title = get_report_title(report_type)
    ws.title = title[:30]  # Excel 工作表名称限制

    # 写入标题行
    headers = get_report_headers(report_type)
    ws.append(headers)

    # 设置标题样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 写入数据
    for row in data:
        ws.append(get_row_data(row, report_type))

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()

def generate_pdf_report(data, report_type):
    """生成 PDF 报表"""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 16)

    # 设置标题
    title = get_report_title(report_type)
    pdf.cell(0, 10, title, 0, 1, 'C')
    pdf.ln(10)

    # 设置表格
    headers = get_report_headers(report_type)
    col_width = pdf.w / (len(headers) + 1)
    row_height = 8

    # 写入表头
    pdf.set_font('Arial', 'B', 10)
    for header in headers:
        pdf.cell(col_width, row_height, header, 1, 0, 'C')
    pdf.ln(row_height)

    # 写入数据
    pdf.set_font('Arial', '', 10)
    for row in data:
        row_data = get_row_data(row, report_type)
        for item in row_data:
            pdf.cell(col_width, row_height, str(item) if item is not None else '', 1, 0, 'L')
        pdf.ln(row_height)

    # 保存到字节流
    output = io.BytesIO()
    pdf.output(dest='S').encode('latin-1')  # PDF 使用 latin-1 编码
    output.write(pdf.output(dest='S').encode('latin-1'))
    output.seek(0)

    return output.getvalue()

def generate_csv_report(data, report_type):
    """生成 CSV 报表"""
    headers = get_report_headers(report_type)

    # 写入 CSV 内容
    csv_content = ','.join(headers) + '\n'
    for row in data:
        row_data = get_row_data(row, report_type)
        # 处理逗号和引号
        processed_row = []
        for item in row_data:
            if item is None:
                processed_row.append('')
            else:
                str_item = str(item)
                if ',' in str_item or '"' in str_item:
                    str_item = str_item.replace('"', '""')
                    processed_row.append(f'"{str_item}"')
                else:
                    processed_row.append(str_item)
        csv_content += ','.join(processed_row) + '\n'

    return csv_content.encode('utf-8-sig')  # UTF-8 带 BOM，支持 Excel

def get_report_title(report_type):
    """获取报表标题"""
    titles = {
        'asset_overview': '资产总览报表',
        'in_use_assets': '使用中资产报表',
        'in_stock_assets': '库存资产报表',
        'transfer_records': '资产领用记录报表',
        'return_records': '资产归还记录报表'
    }
    return titles.get(report_type, '资产报表')

def get_report_headers(report_type):
    """获取报表表头"""
    if report_type in ['asset_overview', 'in_use_assets', 'in_stock_assets']:
        return ['ID', '资产编号', '资产名称', '分类', '状态', '使用人', '部门', '购买日期', '价格', '供应商', '备注']
    elif report_type == 'transfer_records':
        return ['ID', '资产名称', '资产编号', '领用人员', '部门', '领用日期', '领用原因']
    elif report_type == 'return_records':
        return ['ID', '资产名称', '资产编号', '归还人员', '归还日期', '归还状态', '备注']
    else:
        return []

def get_row_data(row, report_type):
    """获取行数据"""
    if report_type in ['asset_overview', 'in_use_assets', 'in_stock_assets']:
        return [
            row.get('id'),
            row.get('asset_number'),
            row.get('name'),
            get_category_text(row.get('category')),
            get_status_text(row.get('status')),
            row.get('user'),
            row.get('department'),
            row.get('purchase_date'),
            f"¥{row.get('price')}" if row.get('price') is not None else '',
            row.get('supplier'),
            row.get('description')
        ]
    elif report_type == 'transfer_records':
        return [
            row.get('id'),
            row.get('asset_name'),
            row.get('asset_number'),
            row.get('user'),
            row.get('department'),
            row.get('transfer_date'),
            row.get('reason')
        ]
    elif report_type == 'return_records':
        return [
            row.get('id'),
            row.get('asset_name'),
            row.get('asset_number'),
            row.get('user'),
            row.get('return_date'),
            get_condition_text(row.get('condition')),
            row.get('notes')
        ]
    else:
        return []

def get_category_text(category):
    """获取分类文本"""
    categories = {
        'office_equipment': '办公设备',
        'electronic_equipment': '电子设备',
        'furniture_appliances': '家具家电',
        'vehicle_equipment': '车辆设备',
        'other': '其他'
    }
    return categories.get(category, category)

def get_status_text(status):
    """获取状态文本"""
    statuses = {
        'in_stock': '库存中',
        'in_use': '使用中',
        'maintenance': '维修中',
        'discarded': '已报废'
    }
    return statuses.get(status, status)

def get_condition_text(condition):
    """获取归还状态文本"""
    conditions = {
        'good': '完好无损',
        'minor_damage': '轻微损坏',
        'major_damage': '严重损坏',
        'unusable': '无法使用'
    }
    return conditions.get(condition, condition)

def main():
    """主函数"""
    try:
        # 从标准输入读取数据
        input_data = sys.stdin.read()
        data = json.loads(input_data)

        report_data = data.get('data', [])
        report_type = data.get('type', 'asset_overview')
        report_format = data.get('format', 'excel')

        # 生成报表
        if report_format == 'excel':
            output = generate_excel_report(report_data, report_type)
        elif report_format == 'pdf':
            output = generate_pdf_report(report_data, report_type)
        elif report_format == 'csv':
            output = generate_csv_report(report_data, report_type)
        else:
            raise ValueError(f'不支持的报表格式: {report_format}')

        # 写入标准输出
        sys.stdout.buffer.write(output)
        sys.stdout.buffer.flush()

    except Exception as e:
        print(f'生成报表失败: {str(e)}', file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()